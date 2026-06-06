"""
DAG standalone de prueba para 1046C - Tesorería Extractos Bancarios BBVA.

Caso de prueba local en Airflow:
- Lee archivos MT940/MT942 desde /opt/airflow/data/input/BBVA
- Lee catálogo desde /opt/airflow/data/baseregistros.xlsx
- Primero enruta por nombre de archivo:
    MULTI -> salida PROMODA
    PRIVA -> salida PRIVALIA
- Para archivos que NO son MULTI ni PRIVA:
    - Separa el archivo por bloques SWIFT usando :20: como inicio de bloque
    - Busca la cuenta bancaria :25: en el catálogo por BankInternalID BBVA
    - Agrupa bloques por SystemID:
        ECC   -> salida ECC
        S4/S4H -> salida S4H
- En BBVA NO existe salida UNMATCHED.
  Si una cuenta no existe en catálogo o tiene SystemID inválido, no se genera archivo para esa cuenta;
  solo se registra en el reporte CSV.
- Mantiene el formato original MT940/MT942.

Salidas por integración:
    /output/bbva/promoda/MT940
    /output/bbva/promoda/MT942
    /output/bbva/privalia/MT940
    /output/bbva/privalia/MT942
    /output/bbva/s4h/MT940
    /output/bbva/s4h/MT942
    /output/bbva/ecc/MT940
    /output/bbva/ecc/MT942

Reportes:
    /output/bbva/_reports

Este DAG es independiente y NO usa el framework general.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from airflow.decorators import dag, task
from airflow.models import Variable

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise ImportError(
        "Falta instalar openpyxl en el ambiente de Airflow. "
        "Instala con: pip install openpyxl"
    ) from exc


# ============================================================
# Configuración standalone de prueba
# ============================================================

DEFAULT_BASE_DIR = "/opt/airflow/data"

BASE_DIR = Path(Variable.get("BBVA_1046C_BASE_DIR", default_var=DEFAULT_BASE_DIR))
INPUT_DIR = Path(Variable.get("BBVA_1046C_INPUT_DIR", default_var=str(BASE_DIR / "input" / "BBVA")))
OUTPUT_DIR = Path(Variable.get("BBVA_1046C_OUTPUT_DIR", default_var=str(BASE_DIR / "output")))
CATALOG_PATH = Path(Variable.get("BBVA_1046C_CATALOG_PATH", default_var=str(BASE_DIR / "baseregistros.xlsx")))

# Con OUTPUT_DIR=/opt/airflow/data/output, el resultado queda:
# /opt/airflow/data/output/bbva/<destino>/<MT940|MT942>
INTEGRATION_OUTPUT_DIR = OUTPUT_DIR / "bbva"

# En BBVA el bank id queda configurable por Variable.
# En Excel puede venir como 12 o 012; se normaliza a 3 dígitos.
BANK_INTERNAL_ID = Variable.get("BBVA_1046C_BANK_INTERNAL_ID", default_var="012")

# Routing por nombre de archivo según NiFi.
PROMODA_FILENAME_TOKEN = "MULTI"
PRIVALIA_FILENAME_TOKEN = "PRIVA"

# Para BBVA normalmente el nombre trae MT940 / MT942 directamente.
MT940_TOKEN = "MT940"
MT942_TOKEN = "MT942"


# ============================================================
# Normalizadores
# ============================================================

def normalize_bank_id(value: object) -> str:
    """
    Normaliza id_bank para comparar:
    12, '12', '012', 12.0 -> '012'
    """
    if value is None:
        return ""

    raw = str(value).strip()

    if raw.endswith(".0"):
        raw = raw[:-2]

    return raw.zfill(3)


def normalize_account(value: object) -> str:
    """
    Normaliza account_number para BBVA y evita fallas por ceros a la izquierda.

    Ejemplo real:
    - En archivo MT940 puede venir :25:0124838610
    - En Excel, si la columna está numérica, se guarda como 124838610

    Para que ambos hagan match, quitamos ceros a la izquierda:
        '0124838610' -> '124838610'
        124838610.0  -> '124838610'
    """
    if value is None:
        return ""

    raw = str(value).strip()

    if raw.endswith(".0"):
        raw = raw[:-2]

    normalized = raw.lstrip("0")
    return normalized or "0"


def normalize_system(value: object) -> Optional[str]:
    """
    Normaliza SystemID para BBVA default:
    S4 o S4H -> S4H
    ECC      -> ECC
    otro     -> None
    """
    if value is None:
        return None

    system = str(value).strip().upper()

    if system in {"S4", "S4H"}:
        return "S4H"

    if system == "ECC":
        return "ECC"

    return None


# ============================================================
# Lectura de catálogo
# ============================================================

def read_catalog(catalog_path: Path, bank_internal_id: str) -> Dict[str, str]:
    """
    Lee baseregistros.xlsx y regresa un diccionario:
        {
            account_number: system_id
        }

    Columnas esperadas:
        id_account | id_bank | account_number | system_id | id_association

    Solo carga registros donde id_bank == BANK_INTERNAL_ID.
    Para BBVA solo se consideran sistemas S4H y ECC.
    """
    if not catalog_path.exists():
        raise FileNotFoundError(f"No existe catálogo: {catalog_path}")

    wb = load_workbook(catalog_path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {
        str(value).strip().lower(): idx
        for idx, value in enumerate(header_row)
        if value is not None
    }

    required_columns = ["id_bank", "account_number", "system_id"]
    missing = [column for column in required_columns if column not in headers]

    if missing:
        raise ValueError(
            f"Faltan columnas requeridas en catálogo {catalog_path}: {missing}"
        )

    expected_bank = normalize_bank_id(bank_internal_id)
    catalog: Dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_bank = normalize_bank_id(row[headers["id_bank"]])

        if row_bank != expected_bank:
            continue

        account = normalize_account(row[headers["account_number"]])
        system = normalize_system(row[headers["system_id"]])

        if account and system:
            catalog[account] = system

    return catalog


# ============================================================
# Lectura / separación MT940-MT942
# ============================================================

def detect_statement_type(filename: str) -> str:
    """
    Detecta MT940 o MT942 usando la nomenclatura del archivo.
    Para BBVA normalmente viene como AXO_MT940_..., MULTI_MT940_..., PRIVA_MT942_...
    """
    upper_name = filename.upper()

    if MT940_TOKEN in upper_name:
        return "MT940"

    if MT942_TOKEN in upper_name:
        return "MT942"

    return "UNKNOWN"


def split_swift_blocks(content: str) -> List[str]:
    """
    Divide el archivo por mensajes/bloques SWIFT.

    Regla:
    - El bloque útil inicia en una línea que empieza con :20:
    - El bloque termina cuando aparece el siguiente :20: o al final del archivo.
    - BBVA puede cerrar el mensaje como '-}' en vez de una línea '-' sola; por eso no dependemos solo del '-'.
    - Se conserva el texto original desde :20: hasta el cierre del bloque.
    """
    blocks: List[str] = []
    current: List[str] = []
    inside_block = False

    for raw_line in content.splitlines():
        line = raw_line.rstrip("\r")

        if line.startswith(":20:"):
            if current:
                blocks.append("\n".join(current).rstrip() + "\n")
                current = []

            inside_block = True
            current.append(line)
            continue

        if inside_block:
            current.append(line)

            # HSBC puede traer '-' solo; BBVA puede traer '-}'.
            # En ambos casos cerramos el bloque.
            if line.strip() in {"-", "-}"}:
                blocks.append("\n".join(current).rstrip() + "\n")
                current = []
                inside_block = False

    if current:
        blocks.append("\n".join(current).rstrip() + "\n")

    return [block for block in blocks if block.strip()]


def extract_ref_and_account_from_block(block: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Obtiene:
    - referencia de mensaje desde :20:
    - cuenta bancaria desde :25:
    """
    reference_20 = None
    account_25 = None

    for line in block.splitlines():
        if line.startswith(":20:"):
            reference_20 = line.replace(":20:", "", 1).strip()

        elif line.startswith(":25:"):
            account_25 = normalize_account(line.replace(":25:", "", 1))

    return reference_20, account_25


# ============================================================
# Routing / Clasificación BBVA
# ============================================================

def detect_filename_destination(filename: str) -> Optional[str]:
    """
    Primer filtro de BBVA por nombre de archivo:
    MULTI -> PROMODA
    PRIVA -> PRIVALIA
    otro  -> None, se procesa por catálogo S4H/ECC
    """
    upper_name = filename.upper()

    if PROMODA_FILENAME_TOKEN in upper_name:
        return "PROMODA"

    if PRIVALIA_FILENAME_TOKEN in upper_name:
        return "PRIVALIA"

    return None


def classify_default_blocks_by_system(
    blocks: List[str],
    catalog: Dict[str, str],
) -> Tuple[Dict[str, List[str]], List[Dict[str, str]]]:
    """
    Clasifica bloques default según cuenta :25: y catálogo.

    Regla importante BBVA:
    - Solo se generan salidas S4H y ECC.
    - Si la cuenta no existe en catálogo, NO hay UNMATCHED.
      Se omite de salida y queda registrado en el reporte.
    """
    grouped: Dict[str, List[str]] = {
        "S4H": [],
        "ECC": [],
    }

    report: List[Dict[str, str]] = []

    for index, block in enumerate(blocks):
        reference_20, account_25 = extract_ref_and_account_from_block(block)

        system = catalog.get(account_25 or "")

        if not system:
            report.append(
                {
                    "block_index": str(index),
                    "reference_20": reference_20 or "",
                    "account_25": account_25 or "",
                    "status": "ACCOUNT_NOT_FOUND_IN_CATALOG_SKIPPED",
                    "resolved_destination": "SKIPPED",
                }
            )
            continue

        system = normalize_system(system)

        if system not in grouped:
            report.append(
                {
                    "block_index": str(index),
                    "reference_20": reference_20 or "",
                    "account_25": account_25 or "",
                    "status": "INVALID_SYSTEM_ID_SKIPPED",
                    "resolved_destination": "SKIPPED",
                }
            )
            continue

        grouped[system].append(block)
        report.append(
            {
                "block_index": str(index),
                "reference_20": reference_20 or "",
                "account_25": account_25 or "",
                "status": "MATCHED",
                "resolved_destination": system,
            }
        )

    return grouped, report


# ============================================================
# Escritura
# ============================================================

def get_output_filename(destination: str, original_filename: str) -> str:
    """
    Regla de nombre:
    - S4H: reemplaza espacios por underscore.
    - Los demás destinos conservan nombre original.
    """
    if destination == "S4H":
        return original_filename.replace(" ", "_")

    return original_filename


def write_text_file(path: Path, content: str) -> None:
    """
    Escribe texto de forma compatible con versiones de Python/Airflow.
    Evita depender de Path.write_text(newline=...), que puede variar por versión.
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", encoding="utf-8", newline="\n") as file:
        file.write(content)


def write_classification_report(report_path: Path, report: List[Dict[str, str]]) -> None:
    """
    Escribe CSV de validación con el destino de cada bloque.
    """
    report_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "block_index",
        "reference_20",
        "account_25",
        "status",
        "resolved_destination",
    ]

    with report_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report)


# ============================================================
# DAG Airflow
# ============================================================

@dag(
    dag_id="prueba_1046c_extractos_bbva_standalone",
    description="Prueba standalone 1046C BBVA: enruta MULTI/PRIVA por filename y AXO/default por catálogo S4H/ECC.",
    start_date=datetime(2025, 1, 1),
    schedule=None,
    catchup=False,
    tags=["1046C", "BBVA", "MT940", "MT942", "standalone", "prueba"],
)
def prueba_1046c_extractos_bbva_standalone():

    @task
    def list_input_files() -> List[str]:
        """
        Lista archivos de entrada.
        Ignora Excel, CSV, tmp, ocultos y reportes.
        """
        INPUT_DIR.mkdir(parents=True, exist_ok=True)

        files = [
            str(path)
            for path in INPUT_DIR.iterdir()
            if path.is_file()
            and not path.name.startswith(".")
            and not path.name.lower().endswith(".classification_report.csv")
            and path.suffix.lower() not in {".xlsx", ".tmp", ".csv"}
        ]

        if not files:
            raise FileNotFoundError(f"No hay archivos para procesar en: {INPUT_DIR}")

        return files

    @task
    def process_file(file_path: str) -> Dict[str, object]:
        """
        Procesa un archivo BBVA MT940/MT942:
        - Si filename contiene MULTI, guarda archivo completo en PROMODA.
        - Si filename contiene PRIVA, guarda archivo completo en PRIVALIA.
        - Si no, divide por bloques y usa catálogo para S4H/ECC.
        - No existe UNMATCHED en BBVA.
        """
        source_path = Path(file_path)
        statement_type = detect_statement_type(source_path.name)
        filename_destination = detect_filename_destination(source_path.name)

        content = source_path.read_text(encoding="utf-8", errors="replace")

        created_files: List[str] = []
        counts_by_destination: Dict[str, int] = {}

        # ====================================================
        # MULTI / PRIVA: no se modifica el archivo.
        # ====================================================
        if filename_destination in {"PROMODA", "PRIVALIA"}:
            output_dir = INTEGRATION_OUTPUT_DIR / filename_destination.lower() / statement_type
            output_filename = get_output_filename(filename_destination, source_path.name)
            output_path = output_dir / output_filename

            write_text_file(output_path, content)
            created_files.append(str(output_path))
            counts_by_destination[filename_destination] = 1

            return {
                "source_file": str(source_path),
                "statement_type": statement_type,
                "routing_mode": "FILENAME",
                "catalog_accounts_loaded": 0,
                "total_blocks": 0,
                "counts_by_destination": counts_by_destination,
                "created_files": created_files,
            }

        # ====================================================
        # DEFAULT: catálogo S4H/ECC, sin UNMATCHED.
        # ====================================================
        catalog = read_catalog(CATALOG_PATH, BANK_INTERNAL_ID)
        blocks = split_swift_blocks(content)

        grouped: Dict[str, List[str]] = {"S4H": [], "ECC": []}
        report: List[Dict[str, str]] = []

        # En BBVA no queremos fallar el DAG si un archivo no trae bloques válidos:
        # se deja registro en reporte y el task termina success.
        if not blocks:
            report.append(
                {
                    "block_index": "",
                    "reference_20": "",
                    "account_25": "",
                    "status": "NO_SWIFT_BLOCKS_FOUND_SKIPPED",
                    "resolved_destination": "SKIPPED",
                }
            )
        else:
            grouped, report = classify_default_blocks_by_system(blocks, catalog)

        for destination, destination_blocks in grouped.items():
            counts_by_destination[destination] = len(destination_blocks)

            if not destination_blocks:
                continue

            output_dir = INTEGRATION_OUTPUT_DIR / destination.lower() / statement_type
            output_filename = get_output_filename(destination, source_path.name)
            output_path = output_dir / output_filename

            # Mantiene el formato original concatenando bloques completos.
            output_content = "".join(destination_blocks)

            write_text_file(output_path, output_content)
            created_files.append(str(output_path))

        report_path = INTEGRATION_OUTPUT_DIR / "_reports" / f"{source_path.name}.classification_report.csv"
        write_classification_report(report_path, report)
        created_files.append(str(report_path))

        return {
            "source_file": str(source_path),
            "statement_type": statement_type,
            "routing_mode": "CATALOG_S4H_ECC",
            "catalog_accounts_loaded": len(catalog),
            "total_blocks": len(blocks),
            "counts_by_destination": counts_by_destination,
            "created_files": created_files,
        }

    @task
    def print_summary(results: List[Dict[str, object]]) -> None:
        """
        Imprime resumen en logs.
        """
        print("========== RESUMEN 1046C BBVA ==========")
        print(f"Directorio de salida integración: {INTEGRATION_OUTPUT_DIR}")

        for result in results:
            if not result:
                print("Resultado vacío recibido; se omite.")
                continue

            print(f"Archivo fuente: {result['source_file']}")
            print(f"Tipo detectado: {result['statement_type']}")
            print(f"Modo de ruteo: {result['routing_mode']}")
            print(f"Cuentas en catálogo cargadas: {result['catalog_accounts_loaded']}")
            print(f"Bloques totales: {result['total_blocks']}")
            print(f"Bloques/archivos por destino: {result['counts_by_destination']}")
            print("Archivos generados:")

            for created in result["created_files"]:
                print(f"  - {created}")

            print("----------------------------------------")

    input_files = list_input_files()
    processed_results = process_file.expand(file_path=input_files)
    print_summary(processed_results)


prueba_1046c_extractos_bbva_standalone()
