"""
DAG standalone de prueba para 1046A - Tesorería Extractos Bancarios HSBC.

Caso de prueba local en Airflow:
- Lee archivos MT940/MT942 desde /opt/airflow/data/input
- Lee catálogo desde /opt/airflow/data/baseregistros.xlsx
- Separa el archivo por bloques SWIFT:
    :20: inicio de bloque
    :25: cuenta bancaria
    -    fin de bloque
- Busca la cuenta bancaria en el catálogo por banco HSBC 021
- Agrupa bloques por SystemID:
    ECC   -> salida ECC
    S4/S4H -> salida S4H
    NOSAP -> salida NOSAP
- Si la cuenta NO existe en catálogo, NO se manda a S4H.
  Se manda a UNMATCHED para revisar, replicando mejor la lógica NiFi:
    LookupAttribute unmatched/failure no entra al route S4H/ECC/NOSAP.
- Mantiene el formato original MT940/MT942.

Cambio aplicado:
- Las salidas ahora se guardan por integración:
    /output/hsbc/s4h/MT940
    /output/hsbc/s4h/MT942
    /output/hsbc/ecc/MT940
    /output/hsbc/ecc/MT942
    /output/hsbc/nosap/MT940
    /output/hsbc/nosap/MT942
    /output/hsbc/unmatched/MT940
    /output/hsbc/unmatched/MT942
- Los reportes se guardan en:
    /output/hsbc/_reports

Este DAG es independiente y NO usa el framework general.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from airflow.decorators import dag, task
from airflow.models import Variable

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise ImportError(
        "Falta instalar openpyxl en el ambiente de Airflow. "
        "Instala con: pip install openpyxl"
    ) from exc


# ============================================================
# Configuración standalone de prueba
# ============================================================

DEFAULT_BASE_DIR = "/opt/airflow/data"

BASE_DIR = Path(Variable.get("HSBC_1046A_BASE_DIR", default_var=DEFAULT_BASE_DIR))
INPUT_DIR = Path(Variable.get(
    "HSBC_1046A_INPUT_DIR",
    default_var=str(BASE_DIR / "input" / "HSBC")
))
OUTPUT_DIR = Path(Variable.get("HSBC_1046A_OUTPUT_DIR", default_var=str(BASE_DIR / "output")))
CATALOG_PATH = Path(Variable.get("HSBC_1046A_CATALOG_PATH", default_var=str(BASE_DIR / "baseregistros.xlsx")))

# Nueva carpeta raíz por integración.
# Con OUTPUT_DIR=/opt/airflow/data/output, el resultado queda:
# /opt/airflow/data/output/hsbc/<sistema>/<MT940|MT942>
INTEGRATION_OUTPUT_DIR = OUTPUT_DIR / "hsbc"

# En requerimiento/NiFi HSBC usa BankInternalID 021.
# En Excel puede venir como 21, por eso se normaliza a 3 dígitos.
BANK_INTERNAL_ID = Variable.get("HSBC_1046A_BANK_INTERNAL_ID", default_var="021")

# Importante:
# En NiFi, si LookupAttribute no encuentra cuenta, sale por unmatched/failure.
# No se debe mandar por default a S4H.
UNMATCHED_SYSTEM_ID = "UNMATCHED"

MT940_MARKER = "LCDE.BLQSAM.OC5MM691"
MT942_MARKER = "LCDE.BLQSAM.NXXMM691"


# ============================================================
# Normalizadores
# ============================================================

def normalize_bank_id(value: object) -> str:
    """
    Normaliza id_bank para comparar:
    21, '21', '021', 21.0 -> '021'
    """
    if value is None:
        return ""

    raw = str(value).strip()

    if raw.endswith(".0"):
        raw = raw[:-2]

    return raw.zfill(3)


def normalize_account(value: object) -> str:
    """
    Normaliza account_number:
    4056515414, '4056515414', 4056515414.0 -> '4056515414'
    """
    if value is None:
        return ""

    raw = str(value).strip()

    if raw.endswith(".0"):
        raw = raw[:-2]

    return raw


def normalize_system(value: object) -> Optional[str]:
    """
    Normaliza SystemID:
    S4 o S4H -> S4H
    ECC      -> ECC
    NOSAP    -> NOSAP
    otro     -> None
    """
    if value is None:
        return None

    system = str(value).strip().upper()

    if system in {"S4", "S4H"}:
        return "S4H"

    if system == "ECC":
        return "ECC"

    if system == "NOSAP":
        return "NOSAP"

    return None


# ============================================================
# Lectura de catálogo
# ============================================================

def read_catalog(catalog_path: Path, bank_internal_id: str) -> Dict[str, str]:
    """
    Lee baseregistros.xlsx y regresa un diccionario:
        {
            account_number: system_id
        }

    Columnas esperadas:
        id_account | id_bank | account_number | system_id | id_association

    Solo carga registros donde id_bank == BANK_INTERNAL_ID.
    """
    if not catalog_path.exists():
        raise FileNotFoundError(f"No existe catálogo: {catalog_path}")

    wb = load_workbook(catalog_path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = {
        str(value).strip().lower(): idx
        for idx, value in enumerate(header_row)
        if value is not None
    }

    required_columns = ["id_bank", "account_number", "system_id"]
    missing = [column for column in required_columns if column not in headers]

    if missing:
        raise ValueError(
            f"Faltan columnas requeridas en catálogo {catalog_path}: {missing}"
        )

    expected_bank = normalize_bank_id(bank_internal_id)
    catalog: Dict[str, str] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_bank = normalize_bank_id(row[headers["id_bank"]])

        if row_bank != expected_bank:
            continue

        account = normalize_account(row[headers["account_number"]])
        system = normalize_system(row[headers["system_id"]])

        if account and system:
            catalog[account] = system

    return catalog


# ============================================================
# Lectura / separación MT940-MT942
# ============================================================

def detect_statement_type(filename: str) -> str:
    """
    Detecta MT940 o MT942 usando la nomenclatura del archivo.
    """
    upper_name = filename.upper()

    if MT940_MARKER in upper_name:
        return "MT940"

    if MT942_MARKER in upper_name:
        return "MT942"

    # Para pruebas con nombres como MT940.txt.
    return "UNKNOWN"


def split_swift_blocks(content: str) -> List[str]:
    """
    Divide el archivo por mensajes/bloques SWIFT.

    Regla:
    - El bloque inicia en una línea que empieza con :20:
    - El bloque termina en una línea que contiene solo -
    - Se conserva el texto original del bloque.
    """
    blocks: List[str] = []
    current: List[str] = []

    for raw_line in content.splitlines():
        # Quitar solo carriage return; conservar espacios útiles de línea.
        line = raw_line.rstrip("\r")

        # Si aparece nuevo :20: y ya traemos bloque abierto, cerramos el anterior.
        # Esto protege contra archivos sin '-' correcto.
        if line.startswith(":20:") and current:
            blocks.append("\n".join(current).rstrip() + "\n")
            current = []

        current.append(line)

        if line.strip() == "-":
            blocks.append("\n".join(current).rstrip() + "\n")
            current = []

    if current:
        blocks.append("\n".join(current).rstrip() + "\n")

    return [block for block in blocks if block.strip()]


def extract_ref_and_account_from_block(block: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Obtiene:
    - referencia de mensaje desde :20:
    - cuenta bancaria desde :25:
    """
    reference_20 = None
    account_25 = None

    for line in block.splitlines():
        if line.startswith(":20:"):
            reference_20 = line.replace(":20:", "", 1).strip()

        elif line.startswith(":25:"):
            account_25 = normalize_account(line.replace(":25:", "", 1))

    return reference_20, account_25


# ============================================================
# Clasificación
# ============================================================

def classify_blocks_by_system(
    blocks: List[str],
    catalog: Dict[str, str],
) -> Tuple[Dict[str, List[str]], List[Dict[str, str]]]:
    """
    Clasifica cada bloque según cuenta :25: y catálogo.

    Salidas:
    - grouped: bloques agrupados por sistema
    - report: detalle por bloque para debugging/validación

    Regla importante:
    Cuenta no encontrada NO va a S4H.
    Va a UNMATCHED para revisión.
    """
    grouped: Dict[str, List[str]] = {
        "S4H": [],
        "ECC": [],
        "NOSAP": [],
        UNMATCHED_SYSTEM_ID: [],
    }

    report: List[Dict[str, str]] = []

    for index, block in enumerate(blocks):
        reference_20, account_25 = extract_ref_and_account_from_block(block)

        system = catalog.get(account_25 or "")

        if not system:
            grouped[UNMATCHED_SYSTEM_ID].append(block)
            report.append(
                {
                    "block_index": str(index),
                    "reference_20": reference_20 or "",
                    "account_25": account_25 or "",
                    "status": "ACCOUNT_NOT_FOUND_IN_CATALOG",
                    "resolved_system": UNMATCHED_SYSTEM_ID,
                }
            )
            continue

        system = normalize_system(system)

        if not system:
            grouped[UNMATCHED_SYSTEM_ID].append(block)
            report.append(
                {
                    "block_index": str(index),
                    "reference_20": reference_20 or "",
                    "account_25": account_25 or "",
                    "status": "INVALID_SYSTEM_ID",
                    "resolved_system": UNMATCHED_SYSTEM_ID,
                }
            )
            continue

        grouped[system].append(block)
        report.append(
            {
                "block_index": str(index),
                "reference_20": reference_20 or "",
                "account_25": account_25 or "",
                "status": "MATCHED",
                "resolved_system": system,
            }
        )

    return grouped, report


# ============================================================
# Escritura
# ============================================================

def get_output_filename(system_id: str, original_filename: str) -> str:
    """
    Regla de nombre:
    - ECC/NOSAP/UNMATCHED: conserva nombre original.
    - S4H: reemplaza espacios por underscore.
    """
    if system_id == "S4H":
        return original_filename.replace(" ", "_")

    return original_filename


def write_classification_report(report_path: Path, report: List[Dict[str, str]]) -> None:
    """
    Escribe CSV de validación con el destino de cada bloque.
    """
    report_path.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "block_index",
        "reference_20",
        "account_25",
        "status",
        "resolved_system",
    ]

    with report_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report)


# ============================================================
# DAG Airflow
# ============================================================

@dag(
    dag_id="prueba_1046a_extractos_hsbc_standalone",
    description="Prueba standalone 1046A HSBC: separa MT940/MT942 por SystemID.",
    start_date=datetime(2025, 1, 1),
    schedule=None,
    catchup=False,
    tags=["1046A", "HSBC", "MT940", "MT942", "standalone", "prueba"],
)
def prueba_1046a_extractos_hsbc_standalone():

    @task
    def list_input_files() -> List[str]:
        """
        Lista archivos de entrada.
        Ignora Excel, CSV, tmp y ocultos.
        """
        INPUT_DIR.mkdir(parents=True, exist_ok=True)

        files = [
            str(path)
            for path in INPUT_DIR.iterdir()
            if path.is_file()
            and not path.name.startswith(".")
            and path.suffix.lower() not in {".xlsx", ".tmp", ".csv"}
        ]

        if not files:
            raise FileNotFoundError(f"No hay archivos para procesar en: {INPUT_DIR}")

        return files

    @task
    def process_file(file_path: str) -> Dict[str, object]:
        """
        Procesa un archivo MT940/MT942:
        - Lee catálogo
        - Divide en bloques
        - Clasifica por SystemID
        - Escribe archivos separados por integración/sistema/tipo
        - Escribe reporte CSV por integración
        """
        source_path = Path(file_path)

        catalog = read_catalog(CATALOG_PATH, BANK_INTERNAL_ID)

        content = source_path.read_text(encoding="utf-8", errors="replace")
        blocks = split_swift_blocks(content)

        if not blocks:
            raise ValueError(f"No se encontraron bloques SWIFT en archivo: {source_path}")

        grouped, report = classify_blocks_by_system(blocks, catalog)

        statement_type = detect_statement_type(source_path.name)

        created_files: List[str] = []
        counts_by_system: Dict[str, int] = {}

        for system_id, system_blocks in grouped.items():
            counts_by_system[system_id] = len(system_blocks)

            if not system_blocks:
                continue

            # CAMBIO PRINCIPAL:
            # Antes:
            #   OUTPUT_DIR / system_id / statement_type
            # Ahora:
            #   OUTPUT_DIR / "hsbc" / system_id.lower() / statement_type
            output_dir = INTEGRATION_OUTPUT_DIR / system_id.lower() / statement_type
            output_dir.mkdir(parents=True, exist_ok=True)

            output_filename = get_output_filename(system_id, source_path.name)
            output_path = output_dir / output_filename

            # Mantiene el formato original concatenando bloques completos.
            output_content = "".join(system_blocks)

            output_path.write_text(output_content, encoding="utf-8", newline="\n")
            created_files.append(str(output_path))

        report_path = INTEGRATION_OUTPUT_DIR / "_reports" / f"{source_path.name}.classification_report.csv"
        write_classification_report(report_path, report)
        created_files.append(str(report_path))

        return {
            "source_file": str(source_path),
            "statement_type": statement_type,
            "catalog_accounts_loaded": len(catalog),
            "total_blocks": len(blocks),
            "counts_by_system": counts_by_system,
            "created_files": created_files,
        }

    @task
    def print_summary(results: List[Dict[str, object]]) -> None:
        """
        Imprime resumen en logs.
        """
        print("========== RESUMEN 1046A HSBC ==========")
        print(f"Directorio de salida integración: {INTEGRATION_OUTPUT_DIR}")

        for result in results:
            print(f"Archivo fuente: {result['source_file']}")
            print(f"Tipo detectado: {result['statement_type']}")
            print(f"Cuentas en catálogo cargadas: {result['catalog_accounts_loaded']}")
            print(f"Bloques totales: {result['total_blocks']}")
            print(f"Bloques por sistema: {result['counts_by_system']}")
            print("Archivos generados:")

            for created in result["created_files"]:
                print(f"  - {created}")

            print("----------------------------------------")

    input_files = list_input_files()
    processed_results = process_file.expand(file_path=input_files)
    print_summary(processed_results)


prueba_1046a_extractos_hsbc_standalone()
