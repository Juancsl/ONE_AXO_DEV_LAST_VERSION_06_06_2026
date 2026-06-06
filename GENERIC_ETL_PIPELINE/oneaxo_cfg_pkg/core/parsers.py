import csv
import json

from io import StringIO, BytesIO
from openpyxl import load_workbook

import xmltodict


def ensure_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def get_path_value(data, path: str):
    if path in (None, "", "."):
        return data

    current = data

    for part in path.split("."):
        if current is None:
            return None

        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None

    return current


def apply_root_path(data, root_path):
    if not root_path:
        return data

    current = data
    for part in root_path:
        if not isinstance(current, dict):
            raise ValueError(f"No se pudo navegar root_path={root_path}; nodo actual no es dict")
        current = current.get(part)
    return current


def parse_xml_bytes(file_bytes: bytes, parser_config: dict):
    encoding = parser_config.get("encoding", "utf-8")
    text = file_bytes.decode(encoding)
    data = xmltodict.parse(text)

    return apply_root_path(data, parser_config.get("root_path"))


def parse_json_bytes(file_bytes: bytes, parser_config: dict):
    encoding = parser_config.get("encoding", "utf-8")
    text = file_bytes.decode(encoding)
    data = json.loads(text)
    root_path = parser_config.get("root_path")
    return apply_root_path(data, root_path) if root_path else data


def parse_csv_bytes(file_bytes: bytes, parser_config: dict):
    delimiter = parser_config.get("delimiter", ",")
    quotechar = parser_config.get("quotechar", '"')
    encoding = parser_config.get("encoding", "utf-8")

    text = file_bytes.decode(encoding)
    reader = csv.DictReader(StringIO(text), delimiter=delimiter, quotechar=quotechar)
    return list(reader)


def parse_xlsx_bytes(file_bytes: bytes, parser_config: dict):
    sheet_name = parser_config.get("sheet_name")
    header_row = int(parser_config.get("header_row", 1))

    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    headers = []
    for cell in ws[header_row]:
        header = "" if cell.value is None else str(cell.value).strip()
        headers.append(header)

    rows = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        record = {}

        for idx, header in enumerate(headers):
            if not header:
                continue

            value = row[idx] if idx < len(row) else None

            if value is None:
                value = ""
            else:
                value = str(value).strip()

            record[header] = value

        if any(str(v).strip() for v in record.values()):
            rows.append(record)

    return rows


def parse_input_bytes(file_bytes: bytes, source_format: str, parser_config: dict):
    source_format = source_format.lower()

    if source_format == "xml":
        return parse_xml_bytes(file_bytes, parser_config)

    if source_format == "json":
        return parse_json_bytes(file_bytes, parser_config)

    if source_format == "csv":
        return parse_csv_bytes(file_bytes, parser_config)

    if source_format in ("xlsx", "excel"):
        return parse_xlsx_bytes(file_bytes, parser_config)

    raise ValueError(f"Formato no soportado: {source_format}")